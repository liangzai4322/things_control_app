from __future__ import annotations

import csv
import json
import shutil
import sys
from collections import OrderedDict, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUTS = ROOT / "outputs"
RUN_NAME = "merged-feishu-expanded-table-" + datetime.now().strftime("%Y%m%d-%H%M%S")
OUT_DIR = OUTPUTS / RUN_NAME
BAD_DIR = OUT_DIR / "bad-json-files"

LIBRARY_ORDER = ["需求库", "流量库", "营销库", "变现库", "案例库", "壁垒库"]
SOURCE_COLUMNS = ["库名", "来源批次", "来源文件", "来源路径"]
CHECK_MARK = "\u2705"


def iter_json_files() -> list[Path]:
    files: list[Path] = []
    for json_dir in OUTPUTS.rglob("json文件"):
        if CHECK_MARK not in str(json_dir):
            continue
        if "merged-feishu" in str(json_dir):
            continue
        files.extend(sorted(json_dir.glob("*.json"), key=lambda p: p.name.lower()))
    return files


def scalar_to_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False)


def add_keys(target: OrderedDict[str, None], row: dict) -> None:
    for key in row.keys():
        target.setdefault(str(key), None)


def normalize_rows(value) -> list[dict]:
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    if isinstance(value, dict):
        return [value]
    return []


def autosize(ws, max_width: int = 60) -> None:
    for col_idx, column_cells in enumerate(ws.columns, 1):
        width = 10
        for cell in column_cells[:80]:
            text = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(text) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws)


def write_csv(path: Path, headers: list[str], rows: list[dict], field_columns: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([row.get(col, "") for col in SOURCE_COLUMNS] + [row["字段"].get(col, "") for col in field_columns])


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    BAD_DIR.mkdir(parents=True, exist_ok=True)

    all_rows: list[dict] = []
    rows_by_library: dict[str, list[dict]] = defaultdict(list)
    fields_by_library: dict[str, OrderedDict[str, None]] = defaultdict(OrderedDict)
    all_fields: OrderedDict[str, None] = OrderedDict()
    bad: list[dict[str, str]] = []
    files = iter_json_files()

    for file_path in files:
        try:
            payload = json.loads(file_path.read_text(encoding="utf-8-sig"))
        except Exception as exc:
            target = BAD_DIR / f"{file_path.parent.parent.name}__{file_path.name}"
            shutil.copy2(file_path, target)
            bad.append({"path": str(file_path), "copied_to": str(target), "error": str(exc)})
            continue

        if not isinstance(payload, dict):
            continue

        batch = file_path.parent.parent.name
        rel = str(file_path.relative_to(ROOT))
        for library in LIBRARY_ORDER:
            for source_row in normalize_rows(payload.get(library, [])):
                row_fields = {str(k): scalar_to_cell(v) for k, v in source_row.items()}
                add_keys(fields_by_library[library], row_fields)
                add_keys(all_fields, row_fields)
                record = {
                    "库名": library,
                    "来源批次": batch,
                    "来源文件": file_path.name,
                    "来源路径": rel,
                    "字段": row_fields,
                }
                rows_by_library[library].append(record)
                all_rows.append(record)

    all_field_columns = list(all_fields.keys())
    all_headers = SOURCE_COLUMNS + all_field_columns

    xlsx_path = OUT_DIR / "merged-json-expanded-table.xlsx"
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "全部展开"
    ws_all.append(all_headers)
    for row in all_rows:
        ws_all.append([row.get(col, "") for col in SOURCE_COLUMNS] + [row["字段"].get(col, "") for col in all_field_columns])
    style_sheet(ws_all)

    for library in LIBRARY_ORDER:
        field_columns = list(fields_by_library[library].keys())
        headers = SOURCE_COLUMNS + field_columns
        ws = wb.create_sheet(library)
        ws.append(headers)
        for row in rows_by_library[library]:
            ws.append([row.get(col, "") for col in SOURCE_COLUMNS] + [row["字段"].get(col, "") for col in field_columns])
        style_sheet(ws)

    wb.save(xlsx_path)

    csv_path = OUT_DIR / "merged-json-expanded-all.csv"
    write_csv(csv_path, all_headers, all_rows, all_field_columns)
    for library in LIBRARY_ORDER:
        field_columns = list(fields_by_library[library].keys())
        write_csv(OUT_DIR / f"{library}.csv", SOURCE_COLUMNS + field_columns, rows_by_library[library], field_columns)

    manifest = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "source_root": str(OUTPUTS),
        "json_file_count": len(files),
        "expanded_row_count": len(all_rows),
        "row_count_by_library": {library: len(rows_by_library[library]) for library in LIBRARY_ORDER},
        "bad_json_count": len(bad),
        "xlsx_path": str(xlsx_path),
        "csv_path": str(csv_path),
        "bad_json_dir": str(BAD_DIR),
        "bad_json_files": bad,
    }
    manifest_path = OUT_DIR / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
