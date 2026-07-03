from __future__ import annotations

import csv
import json
import shutil
import sys
from collections import OrderedDict, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "outputs" / "汇总"
RUN_DIR = SOURCE_DIR / ("展开合并表-" + datetime.now().strftime("%Y%m%d-%H%M%S"))
BAD_DIR = RUN_DIR / "语法或结构问题json"

LIBRARIES = ["需求库", "流量库", "营销库", "变现库", "案例库", "壁垒库"]
SOURCE_COLUMNS = ["库名", "来源文件"]


def scalar_to_cell(value):
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False)


def add_keys(target: OrderedDict[str, None], row: dict) -> None:
    for key in row.keys():
        target.setdefault(str(key), None)


def validate_payload(payload) -> str | None:
    if not isinstance(payload, dict):
        return "顶层不是对象"

    keys = list(payload.keys())
    missing = [key for key in LIBRARIES if key not in payload]
    extra = [key for key in keys if key not in LIBRARIES]
    if missing:
        return "缺少键: " + ",".join(missing)
    if extra:
        return "存在额外键: " + ",".join(extra)

    for key in LIBRARIES:
        if not isinstance(payload[key], list):
            return f"{key} 的值不是数组"
        for index, item in enumerate(payload[key], start=1):
            if not isinstance(item, dict):
                return f"{key} 第 {index} 行不是对象"
    return None


def autosize(ws, max_width: int = 64) -> None:
    for col_idx, cells in enumerate(ws.columns, start=1):
        width = 10
        for cell in list(cells)[:100]:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(max_width, len(value) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def style_sheet(ws) -> None:
    fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    autosize(ws)


def append_rows(ws, headers: list[str], rows: list[dict], field_columns: list[str]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(col, "") for col in SOURCE_COLUMNS] + [row["字段"].get(col, "") for col in field_columns])
    style_sheet(ws)


def write_csv(path: Path, headers: list[str], rows: list[dict], field_columns: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([row.get(col, "") for col in SOURCE_COLUMNS] + [row["字段"].get(col, "") for col in field_columns])


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    RUN_DIR.mkdir(parents=True, exist_ok=True)
    BAD_DIR.mkdir(parents=True, exist_ok=True)

    files = sorted(SOURCE_DIR.glob("*.json"), key=lambda path: path.name.lower())
    all_rows: list[dict] = []
    rows_by_library: dict[str, list[dict]] = defaultdict(list)
    all_fields: OrderedDict[str, None] = OrderedDict()
    fields_by_library: dict[str, OrderedDict[str, None]] = defaultdict(OrderedDict)
    skipped: list[dict[str, str]] = []

    for file_path in files:
        try:
            payload = json.loads(file_path.read_text(encoding="utf-8-sig"))
        except Exception as exc:
            dest = BAD_DIR / file_path.name
            shutil.copy2(file_path, dest)
            skipped.append({"file": file_path.name, "reason": "JSON解析失败: " + str(exc), "copied_to": str(dest)})
            continue

        reason = validate_payload(payload)
        if reason:
            dest = BAD_DIR / file_path.name
            shutil.copy2(file_path, dest)
            skipped.append({"file": file_path.name, "reason": reason, "copied_to": str(dest)})
            continue

        for library in LIBRARIES:
            for source_row in payload[library]:
                row_fields = {str(k): scalar_to_cell(v) for k, v in source_row.items()}
                add_keys(all_fields, row_fields)
                add_keys(fields_by_library[library], row_fields)
                row = {"库名": library, "来源文件": file_path.name, "字段": row_fields}
                all_rows.append(row)
                rows_by_library[library].append(row)

    wb = Workbook()
    all_field_columns = list(all_fields.keys())
    all_headers = SOURCE_COLUMNS + all_field_columns
    ws_all = wb.active
    ws_all.title = "全部展开"
    append_rows(ws_all, all_headers, all_rows, all_field_columns)

    for library in LIBRARIES:
        field_columns = list(fields_by_library[library].keys())
        headers = SOURCE_COLUMNS + field_columns
        ws = wb.create_sheet(library)
        append_rows(ws, headers, rows_by_library[library], field_columns)

    xlsx_path = RUN_DIR / "汇总-json-展开合并表.xlsx"
    wb.save(xlsx_path)

    csv_path = RUN_DIR / "汇总-json-全部展开.csv"
    write_csv(csv_path, all_headers, all_rows, all_field_columns)
    for library in LIBRARIES:
        field_columns = list(fields_by_library[library].keys())
        write_csv(RUN_DIR / f"{library}.csv", SOURCE_COLUMNS + field_columns, rows_by_library[library], field_columns)

    manifest = {
        "source_dir": str(SOURCE_DIR),
        "run_dir": str(RUN_DIR),
        "checked_json_count": len(files),
        "merged_json_count": len(files) - len(skipped),
        "skipped_json_count": len(skipped),
        "expanded_row_count": len(all_rows),
        "row_count_by_library": {library: len(rows_by_library[library]) for library in LIBRARIES},
        "xlsx_path": str(xlsx_path),
        "csv_path": str(csv_path),
        "bad_json_dir": str(BAD_DIR),
        "skipped_files": skipped,
    }
    manifest_path = RUN_DIR / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
