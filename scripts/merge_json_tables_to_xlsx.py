from __future__ import annotations

import csv
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUTS = ROOT / "outputs"
RUN_NAME = "merged-feishu-table-" + datetime.now().strftime("%Y%m%d-%H%M%S")
OUT_DIR = OUTPUTS / RUN_NAME
BAD_DIR = OUT_DIR / "bad-json-files"

MAIN_COLUMNS = ["需求库", "流量库", "营销库", "变现库", "案例库", "壁垒库"]
SOURCE_COLUMNS = ["来源批次", "来源文件", "来源路径"]
HEADERS = SOURCE_COLUMNS + MAIN_COLUMNS


def iter_json_files() -> list[Path]:
    files: list[Path] = []
    for json_dir in OUTPUTS.rglob("json文件"):
        if "✅" not in str(json_dir):
            continue
        files.extend(sorted(json_dir.glob("*.json"), key=lambda p: p.name.lower()))
    return files


def as_rows(payload):
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("data", "rows"):
            if isinstance(payload.get(key), list):
                return payload[key]
        return [payload]
    return []


def scalar_to_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return str(value)
    return json.dumps(value, ensure_ascii=False)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    BAD_DIR.mkdir(parents=True, exist_ok=True)

    rows: list[list[str]] = []
    bad: list[dict[str, str]] = []
    files = iter_json_files()

    for file_path in files:
        try:
            payload = json.loads(file_path.read_text(encoding="utf-8-sig"))
        except Exception as exc:
            target = BAD_DIR / f"{file_path.parent.parent.name}__{file_path.name}"
            shutil.copy2(file_path, target)
            bad.append(
                {
                    "path": str(file_path),
                    "copied_to": str(target),
                    "error": str(exc),
                }
            )
            continue

        batch = file_path.parent.parent.name
        rel = file_path.relative_to(ROOT)
        for item in as_rows(payload):
            if not isinstance(item, dict):
                continue
            rows.append(
                [
                    batch,
                    file_path.name,
                    str(rel),
                    *[scalar_to_cell(item.get(col, "")) for col in MAIN_COLUMNS],
                ]
            )

    csv_path = OUT_DIR / "merged-json-table.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(HEADERS)
        writer.writerows(rows)

    xlsx_path = OUT_DIR / "merged-json-table.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "合并表"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 26,
        "B": 18,
        "C": 64,
        "D": 50,
        "E": 50,
        "F": 50,
        "G": 50,
        "H": 50,
        "I": 50,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(xlsx_path)

    manifest = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "source_root": str(OUTPUTS),
        "json_file_count": len(files),
        "merged_row_count": len(rows),
        "bad_json_count": len(bad),
        "columns": HEADERS,
        "xlsx_path": str(xlsx_path),
        "csv_path": str(csv_path),
        "bad_json_dir": str(BAD_DIR),
        "bad_json_files": bad,
    }
    manifest_path = OUT_DIR / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    main()
